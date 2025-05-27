#!/usr/bin/env python3
"""
Script to sample matching entries from multiple JSONL files based on shared hash field.
"""
import argparse
import json
import random
import csv
import sys
from pathlib import Path
from collections import defaultdict
from typing import List, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def find_jsonl_files(folder_path: str) -> List[str]:
    """Find all JSONL files in the given folder."""
    folder = Path(folder_path)
    if not folder.exists():
        print(f"Error: Folder {folder_path} does not exist.")
        sys.exit(1)
    
    if not folder.is_dir():
        print(f"Error: {folder_path} is not a directory.")
        sys.exit(1)
    
    # Find all .jsonl files
    jsonl_files = list(folder.glob("*.jsonl"))
    
    if not jsonl_files:
        print(f"Error: No JSONL files found in {folder_path}")
        sys.exit(1)
    
    # Convert to strings and sort for consistent ordering
    return sorted(str(f) for f in jsonl_files)

def read_jsonl_file(filepath: str) -> List[Dict[Any, Any]]:
    """Read a JSONL file and return list of JSON objects."""
    entries = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if line:
                    try:
                        entries.append(json.loads(line))
                    except json.JSONDecodeError as e:
                        print(f"Warning: Invalid JSON on line {line_num} in {filepath}: {e}")
    except FileNotFoundError:
        print(f"Error: File {filepath} not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        sys.exit(1)
    
    return entries

def group_entries_by_hash(files: List[str], hash_field: str = "prompt_hash") -> Dict[str, List[Dict]]:
    """
    Group entries from multiple JSONL files by their hash field.
    
    Args:
        files: List of JSONL file paths
        hash_field: Name of the hash field to group by
        
    Returns:
        Dictionary mapping hash values to lists of entries with that hash
    """
    hash_groups = defaultdict(list)
    
    for filepath in files:
        print(f"Processing {filepath}...")
        entries = read_jsonl_file(filepath)
        
        for entry in entries:
            if hash_field in entry:
                hash_value = entry[hash_field]
                # Add source file info to the entry
                entry_with_source = entry.copy()
                entry_with_source['_source_file'] = Path(filepath).name
                hash_groups[hash_value].append(entry_with_source)
            else:
                print(f"Warning: Entry missing '{hash_field}' field in {filepath}")
    
    return hash_groups

def sample_entries(hash_groups: Dict[str, List[Dict]], num_samples: int = 10) -> List[Dict]:
    """
    Randomly sample entries from hash groups.
    
    Args:
        hash_groups: Dictionary of hash -> list of entries
        num_samples: Number of hash groups to sample
        
    Returns:
        List of sampled entries
    """
    # Filter to only include hashes that have entries from multiple files
    multi_file_hashes = {
        hash_val: entries for hash_val, entries in hash_groups.items()
        if len(set(entry['_source_file'] for entry in entries)) > 1
    }
    
    if not multi_file_hashes:
        print("Warning: No hash values found across multiple files.")
        multi_file_hashes = hash_groups
    
    # Sample hash groups
    available_hashes = list(multi_file_hashes.keys())
    num_to_sample = min(num_samples, len(available_hashes))
    
    if num_to_sample < num_samples:
        print(f"Warning: Only {num_to_sample} hash groups available, sampling all of them.")
    
    selected_hashes = random.sample(available_hashes, num_to_sample)
    
    # Collect all entries from selected hash groups
    sampled_entries = []
    for hash_val in selected_hashes:
        entries = multi_file_hashes[hash_val]
        sampled_entries.extend(entries)
    
    return sampled_entries

def write_to_xlsx(entries: List[Dict], output_file: str):
    """Write entries to XLSX file."""
    if not entries:
        print("No entries to write.")
        return
    
    # Define the columns we want to keep (excluding evaluation, prompt, and prompt_hash)
    desired_columns = ['fable', 'translated_fable', 'translation_model']
    
    # Define score types to extract from evaluation
    score_types = ['accuracy', 'fluency', 'coherence', 'style', 'cultural_pragmatic']
    score_columns = [f'{score_type}_score' for score_type in score_types]
    
    # Get all unique field names across all entries
    all_fields = set()
    for entry in entries:
        all_fields.update(entry.keys())
    
    # Filter to only include desired columns that exist in the data
    fieldnames = [col for col in desired_columns if col in all_fields]
    
    # Add score columns and average
    fieldnames.extend(score_columns)
    fieldnames.append('average_score')
    
    # Add _source_file for reference
    if '_source_file' in all_fields:
        fieldnames.append('_source_file')
    
    try:
        # Group entries by prompt_hash (still use for grouping even though not in output)
        grouped_entries = defaultdict(list)
        for entry in entries:
            prompt_hash = entry.get('prompt_hash', 'unknown')
            grouped_entries[prompt_hash].append(entry)
        
        # Prepare data for DataFrame with blank rows between groups
        rows = []
        group_keys = sorted(grouped_entries.keys())
        
        for i, prompt_hash in enumerate(group_keys):
            group_entries = grouped_entries[prompt_hash]
            
            # Process each entry in the group
            for entry in group_entries:
                row = {}
                
                # Handle regular fields
                for field in desired_columns:
                    if field in entry:
                        value = entry[field]
                        if isinstance(value, (dict, list)):
                            row[field] = json.dumps(value, ensure_ascii=False)
                        else:
                            row[field] = value
                    else:
                        row[field] = None
                
                # Extract scores from evaluation field
                scores = []
                if 'evaluation' in entry:
                    try:
                        eval_value = entry['evaluation']
                        # Handle both string and dict cases
                        if isinstance(eval_value, str):
                            eval_data = json.loads(eval_value)
                        elif isinstance(eval_value, dict):
                            eval_data = eval_value
                        else:
                            eval_data = {}
                        
                        for score_type in score_types:
                            if score_type in eval_data:
                                score_obj = eval_data[score_type]
                                if isinstance(score_obj, dict) and 'score' in score_obj:
                                    score = score_obj['score']
                                    row[f'{score_type}_score'] = score
                                    scores.append(score)
                                elif isinstance(score_obj, (int, float)):
                                    # Direct score value
                                    row[f'{score_type}_score'] = score_obj
                                    scores.append(score_obj)
                                else:
                                    row[f'{score_type}_score'] = None
                            else:
                                row[f'{score_type}_score'] = None
                    except Exception as e:
                        print(f"Warning: Failed to parse evaluation for entry: {e}")
                        # If parsing fails, set all scores to None
                        for score_type in score_types:
                            row[f'{score_type}_score'] = None
                else:
                    # No evaluation field, set all scores to None
                    for score_type in score_types:
                        row[f'{score_type}_score'] = None
                
                # Calculate average score
                if scores:
                    row['average_score'] = round(sum(scores) / len(scores), 2)
                else:
                    row['average_score'] = None
                
                # Add source file
                if '_source_file' in entry:
                    row['_source_file'] = entry['_source_file']
                else:
                    row['_source_file'] = None
                        
                rows.append(row)
            
            # Add blank row between groups (except after the last group)
            if i < len(group_keys) - 1:
                blank_row = {field: None for field in fieldnames}
                rows.append(blank_row)
        
        # Create DataFrame and write to Excel
        df = pd.DataFrame(rows, columns=fieldnames)
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Adjust column widths
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Set wider columns for specific fields
        column_widths = {
            'fable': 60,
            'translated_fable': 60,
            'translation_model': 20,
            '_source_file': 25
        }
        
        # Apply column widths
        for col_idx, col_name in enumerate(fieldnames, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            if col_name in column_widths:
                ws.column_dimensions[col_letter].width = column_widths[col_name]
            else:
                # Default width for score columns
                ws.column_dimensions[col_letter].width = 15
        
        # Enable text wrapping for text columns
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        for row in ws.iter_rows(min_row=2):  # Skip header row
            for cell in row:
                col_name = fieldnames[cell.column - 1]
                if col_name in ['fable', 'translated_fable']:
                    cell.alignment = wrap_alignment
        
        wb.save(output_file)
        
        print(f"Successfully wrote {len(entries)} entries to {output_file}")
        print(f"Organized into {len(group_keys)} groups with blank rows between groups")
        print(f"Columns included: {', '.join(fieldnames)}")
        print(f"Extracted individual scores: {', '.join(score_columns)}")
        
    except Exception as e:
        print(f"Error writing to XLSX: {e}")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(
        description="Sample matching entries from multiple JSONL files based on shared hash field"
    )
    parser.add_argument(
        "folder",
        help="Folder containing JSONL files to process"
    )
    parser.add_argument(
        "--output",
        "-o",
        default="sampled_entries.xlsx",
        help="Output XLSX file (default: sampled_entries.xlsx)"
    )
    parser.add_argument(
        "--num-samples",
        "-n",
        type=int,
        default=10,
        help="Number of hash groups to sample (default: 10)"
    )
    parser.add_argument(
        "--hash-field",
        default="prompt_hash",
        help="Name of the hash field to group by (default: hash)"
    )
    parser.add_argument(
        "--seed",
        type=int,
        help="Random seed for reproducible sampling"
    )
    
    args = parser.parse_args()
    
    # Set random seed if provided
    if args.seed:
        random.seed(args.seed)
        print(f"Using random seed: {args.seed}")
    
    # Find all JSONL files in the folder
    jsonl_files = find_jsonl_files(args.folder)
    
    print(f"Found {len(jsonl_files)} JSONL files in {args.folder}:")
    for file in jsonl_files:
        print(f"  - {Path(file).name}")
    print()
    
    print(f"Grouping by field: {args.hash_field}")
    print(f"Sampling {args.num_samples} hash groups")
    print(f"Output file: {args.output}")
    print()
    
    # Group entries by hash
    hash_groups = group_entries_by_hash(jsonl_files, args.hash_field)
    
    print(f"Found {len(hash_groups)} unique hash groups")
    
    # Sample entries
    sampled_entries = sample_entries(hash_groups, args.num_samples)
    
    print(f"Sampled {len(sampled_entries)} total entries")
    
    # Write to XLSX
    write_to_xlsx(sampled_entries, args.output)

if __name__ == "__main__":
    main() 